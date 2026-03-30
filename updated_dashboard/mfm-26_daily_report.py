# импортируем библиотеки
import pandas as pd
from datetime import datetime, date

# словари макрорегионов и федеральных округов + утилиты нормализации
from macroregions import COUNTRY_MACRO_REGIONS
from russian_fo import RUSSIAN_REGIONS

events_name = f"events {datetime.now().strftime('%Y-%m-%d %H-%M')}"
essay_name = f"essay {datetime.now().strftime('%Y-%m-%d %H-%M')}"
video_name = f"video {datetime.now().strftime('%Y-%m-%d %H-%M')}"

print(events_name)
print(essay_name)
print(video_name)

try:
    df = pd.read_csv(f'../data/users.csv', sep=';')
except:
    df = pd.read_csv(f'../data/users.csv', sep=';')

col = [
    'ID',
    'Email #email',
    'Гражданство',
    'Страна проживания',
    'Регион проживания',
    'Возраст',
    'Состояние',
    'Готов добраться за свой счёт #travelownexpense',
    'Участие в РП (да/нет) #expeditionprogram',
    'Дата создания заявки',
    'Имя на русском #firstName.ru',
    'Имя на английском #firstName.en',
    'Нет имени',
    'Фамилия на русском #lastName.ru',
    'Фамилия на английском #lastName.en',
    'Нет фамилии',
    'Отчество на русском #patronymic.ru',
    'Отчество на английском #patronymic.en',
    'Нет отчества',
    'Пол #sex',
    'Дата рождения',
    'Адрес фактического места проживания #addressresidence',
    'Родной язык #nativelanguage',
    'Род деятельность #statusactivity',
    'Направление деятельности #fieldwork',
    'Номер телефона #phone',
    'Фотография #photo',
    'Город фактического места проживания #cityresidence',
    'Язык переписки #correspondencelanguage',
    'Тип питания #dietype',
    'Уровень владения английским языком #englevel',
    'Уровень владения русским языком #ruslevel'
    ]

df = df[col].copy()

try:
    df_video = pd.read_csv(f'../data/video.csv', sep=';')
    df_essay = pd.read_csv(f'../data/esse.csv', sep=';')
except:
    df_video = pd.read_csv(f'../data/video.csv', sep=';')
    df_essay = pd.read_csv(f'../data/esse.csv', sep=';')

# Фильтрация email с @wyffest.com
df = df[~df['Email #email'].str.contains('@wyffest.com', case=False, na=False)]
# Убираем отозванные заявки
df = df[(df['Состояние'] == "Черновик") | (df['Состояние'] == "На рассмотрении")]
df = df[df['Страна проживания'].notna()].copy()

# Подготовка данных для основной статистики
df['is_child'] = df['Возраст'].between(14, 17)
df['is_rf_citizenship'] = df['Гражданство'] == 'Россия'
df['is_rf_residence'] = df['Страна проживания'] == 'Россия'

# Фильтрация эссе и видео по статусу "Отправлена"
VALID_STATUSES = {'Отправлена', 'Оценена', 'На оценке'}
df_essay_filtered = df_essay[df_essay['Состояние'].isin(VALID_STATUSES)].copy()
df_video_filtered = df_video[df_video['Состояние'].isin(VALID_STATUSES)].copy()

# Убираем @wyffest.com из эссе и видео
df_essay_filtered = df_essay_filtered[~df_essay_filtered['Автор'].str.contains('@wyffest.com', case=False, na=False)]
df_video_filtered = df_video_filtered[~df_video_filtered['Автор'].str.contains('@wyffest.com', case=False, na=False)]

# Оставляем только уникальных авторов
df_essay_unique = df_essay_filtered.drop_duplicates(subset=['Автор'])
df_video_unique = df_video_filtered.drop_duplicates(subset=['Автор'])

# Присоединяем к уникальной заявке статус эссе и статус видеовизитки.
# Если эссе или видео нет, то присоединится NaN.
df_merged = df.drop_duplicates(subset=['Email #email']).merge(
    df_essay_unique[['Автор', 'Состояние']].rename(columns={'Автор': 'Email #email', 'Состояние': 'Статус эссе'}),
    on='Email #email',
    how='left'
).merge(
    df_video_unique[['Автор', 'Состояние']].rename(columns={'Автор': 'Email #email', 'Состояние': 'Статус видеовизитки'}),
    on='Email #email',
    how='left'
)

# Статус заявки нам нужен из основной df (там это "Состояние"), поэтому добавим явно:
df_merged['Статус заявки'] = df_merged['Состояние']
df_merged = df_merged.drop(columns=['Состояние'])

# Общая статистика
total = len(df_merged)
children_14_17 = df_merged['is_child'].sum()
adults_18_35 = total - children_14_17

# Общая статистика по эссе и видео
total_essay = len(df_merged[df_merged['Статус эссе'].notna()])
essay_children = df_merged[df_merged['Статус эссе'].notna() & df_merged['is_child']].shape[0]
essay_adults = total_essay - essay_children

total_video = len(df_merged[df_merged['Статус видеовизитки'].notna()])
video_children = df_merged[df_merged['Статус видеовизитки'].notna() & df_merged['is_child']].shape[0]
video_adults = total_video - video_children

# РФ
rf = df_merged[(df_merged['is_rf_citizenship']) & (df_merged['is_rf_residence'])]
rf_total = len(rf)
rf_children = rf['is_child'].sum()
rf_essay = len(rf[rf['Статус эссе'].notna()])
rf_video = len(rf[rf['Статус видеовизитки'].notna()])
rf_children_essay = len(rf[(rf['Статус эссе'].notna()) & (rf['is_child'])])
rf_children_video = len(rf[(rf['Статус видеовизитки'].notna()) & (rf['is_child'])])

# ИН
foreign = df_merged[~df_merged['is_rf_citizenship'] & ~df_merged['is_rf_residence']]
foreign_total = len(foreign)
foreign_children = foreign['is_child'].sum()
foreign_essay = len(foreign[foreign['Статус эссе'].notna()])
foreign_video = len(foreign[foreign['Статус видеовизитки'].notna()])
foreign_children_essay = len(foreign[(foreign['Статус эссе'].notna()) & (foreign['is_child'])])
foreign_children_video = len(foreign[(foreign['Статус видеовизитки'].notna()) & (foreign['is_child'])])

# ИН в РФ
foreign_in_rf = df_merged[~df_merged['is_rf_citizenship'] & df_merged['is_rf_residence']]
foreign_in_rf_total = len(foreign_in_rf)
foreign_in_rf_children = foreign_in_rf['is_child'].sum()
foreign_in_rf_essay = len(foreign_in_rf[foreign_in_rf['Статус эссе'].notna()])
foreign_in_rf_video = len(foreign_in_rf[foreign_in_rf['Статус видеовизитки'].notna()])
foreign_in_rf_children_essay = len(foreign_in_rf[(foreign_in_rf['Статус эссе'].notna()) & (foreign_in_rf['is_child'])])
foreign_in_rf_children_video = len(foreign_in_rf[(foreign_in_rf['Статус видеовизитки'].notna()) & (foreign_in_rf['is_child'])])

# Соотечественники
compatriots = df_merged[df_merged['is_rf_citizenship'] & ~df_merged['is_rf_residence']]
compatriots_total = len(compatriots)
compatriots_children = compatriots['is_child'].sum()
compatriots_essay = len(compatriots[compatriots['Статус эссе'].notna()])
compatriots_video = len(compatriots[compatriots['Статус видеовизитки'].notna()])
compatriots_children_essay = len(compatriots[(compatriots['Статус эссе'].notna()) & (compatriots['is_child'])])
compatriots_children_video = len(compatriots[(compatriots['Статус видеовизитки'].notna()) & (compatriots['is_child'])])

# Количество стран
countries_count = df_merged['Страна проживания'].nunique()

df_merged['Макрорегион'] = df_merged['Страна проживания'].map(COUNTRY_MACRO_REGIONS).fillna('Нет макро')
df_merged['Федеральный округ'] = df_merged['Регион проживания'].map(RUSSIAN_REGIONS).fillna('Нет ФО')

def category(age):
    if 14 <= age <= 17:
        return '14-17'
    elif 18 <= age <= 35:
        return '18-35'
    else:
        return 'Другая категория'  # или 'Не подходит'

df_merged['Категория участника'] = df_merged['Возраст'].apply(category)

df_merged['Дата создания заявки'] = pd.to_datetime(df_merged['Дата создания заявки'], errors='coerce').dt.strftime("%d.%m.%Y")

def is_profile_complete(row):
    # Список обязательных полей из required_fields.py (3-22)
    required_fields = [
        'ID',
        'Email #email',
        'Гражданство',
        'Страна проживания',
        'Пол #sex',
        'Дата рождения',
        'Дата рождения',
        'Номер телефона #phone',
        'Фотография #photo',
        'Город фактического места проживания #cityresidence',
        'Адрес фактического места проживания #addressresidence',
        'Родной язык #nativelanguage',
        'Язык переписки #correspondencelanguage',
        'Тип питания #dietype',
        'Род деятельность #statusactivity',
        'Направление деятельности #fieldwork',
        'Уровень владения английским языком #englevel',
        'Уровень владения русским языком #ruslevel'
    ]
    # Группы "или-или" для ФИО
    either_or_groups = [
        (["Имя на русском #firstName.ru", "Имя на английском #firstName.en", "Нет имени"], "Нет имени"),
        (["Фамилия на русском #lastName.ru", "Фамилия на английском #lastName.en", "Нет фамилии"], "Нет фамилии"),
        (["Отчество на русском #patronymic.ru", "Отчество на английском #patronymic.en", "Нет отчества"], "Нет отчества"),
    ]
    # Проверяем обязательные поля
    for field in required_fields:
        if pd.isna(row.get(field)) or str(row.get(field)).strip() == '':
            return False
    # Проверяем либо хотя бы одно из имен (на русском/английском), либо "Нет имени" == "Да"
    for fields, neg_field in either_or_groups:
        # индексы 0,1 - это имя/фамилия/отчество, 2 - "Нет имя" и т.д.
        if (
            (pd.isna(row.get(fields[0])) or str(row.get(fields[0])).strip() == '') and
            (pd.isna(row.get(fields[1])) or str(row.get(fields[1])).strip() == '')
        ):
            # Если оба пустые, то проверяем "Нет ..."
            if str(row.get(fields[2])).strip() != 'Да':
                return False
    return True

df_merged['Профиль заполнен полностью'] = df_merged.apply(is_profile_complete, axis=1)

df_merged[(df_merged['Страна проживания'] == "Венесуэла") & (df_merged['Возраст'].between(14,17))]

df_merged[[
    'ID',
    'Гражданство',
    'Страна проживания',
    'Регион проживания',
    'Возраст',
    'Статус заявки',
    'is_child',
    'is_rf_citizenship',
    'is_rf_residence',
    'Готов добраться за свой счёт #travelownexpense',
    'Участие в РП (да/нет) #expeditionprogram',
    'Макрорегион',
    'Федеральный округ',
    'Статус эссе',
    'Статус видеовизитки',
    'Категория участника',
    'Дата создания заявки',
    'Профиль заполнен полностью'
]].to_excel(
    r"data/Выгрузка данных от {0}.xlsx".format(date.today()),
    index=False
)

print("Уникальные значения is_rf_residence:", df_merged['is_rf_residence'].unique())

print("Всего с Россией в стране проживания:",
      df_merged[df_merged['Страна проживания'] == 'Россия']['Email #email'].nunique())
print("Из них На рассмотрении:",
      df_merged[(df_merged['Страна проживания'] == 'Россия') &
                (df_merged['Статус заявки'] == 'На рассмотрении')]['Email #email'].nunique())
print("Из них Черновик:",
      df_merged[(df_merged['Страна проживания'] == 'Россия') &
                (df_merged['Статус заявки'] == 'Черновик')]['Email #email'].nunique())


import openpyxl
wb = openpyxl.load_workbook(f'data/Выгрузка данных от {date.today()}.xlsx')
ws = wb.active
# Найдем колонку is_rf_residence
headers = [cell.value for cell in ws[1]]
col_idx = headers.index('is_rf_residence') + 1
# Покажем первые 5 значений и их тип
for row in ws.iter_rows(min_row=2, max_row=6, min_col=col_idx, max_col=col_idx):
    for cell in row:
        print(f"value={cell.value!r}, type={type(cell.value)}")

fo_mapped = df_merged[df_merged['Страна проживания'] == 'Россия']['Федеральный округ'].value_counts()
print(fo_mapped)
print(df_merged[(df_merged['Страна проживания'] == 'Россия') &
                (df_merged['Федеральный округ'] == 'Нет ФО')]['Регион проживания'].unique())

print(df_merged[(df_merged['Страна проживания'] == 'Россия') &
                (df_merged['Федеральный округ'] == 'Нет ФО')][['ID', 'Email #email', 'Регион проживания', 'Гражданство']])